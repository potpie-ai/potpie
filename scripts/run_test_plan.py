#!/usr/bin/env python3
"""Run the QA test plan from Tests.xlsx and write results to a dated sheet.

What it does
------------
1. (Optional) Switches the target repo to a given branch (default: ``main``)
   so all tests execute against that branch.
2. Reads the ``Daily Plan`` sheet from the workbook.
3. For every test row it extracts the shell command(s) embedded in backticks in
   the "Step / Command" column and decides whether the test is auto-runnable.
   Only a conservative allow-list of ``potpie`` CLI commands is executed; rows
   that describe manual/process work (PyPI submission, approvals, dashboards,
   cross-Python-version installs, ...) are recorded as MANUAL and skipped.
4. Runnable commands are executed one by one. stdout/stderr/exit-code are
   captured as evidence and PASS/FAIL is derived from the exit code.
5. Results are printed to the console and written back to the workbook in a new
   sheet whose name is the run date (e.g. ``2026-06-27``).

Usage
-----
    python scripts/run_test_plan.py                 # full run against main
    python scripts/run_test_plan.py --dry-run       # classify only, run nothing
    python scripts/run_test_plan.py --no-checkout   # don't touch git
    python scripts/run_test_plan.py --excel /path/to/Tests.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
from pathlib import Path
import re
import shlex
import subprocess
import sys
import tempfile
import uuid

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")


DEFAULT_REPO = str(Path.cwd())
DEFAULT_EXCEL = str(Path.cwd() / "Context Engine Test Plan for v2 release.xlsx")
SOURCE_SHEET = "Daily Plan"
HEADER_ROW = 4  # row that holds column titles in the source sheet
LOCAL_CLI = ""

# Commands we are willing to execute automatically. Anything that does not start
# with one of these prefixes is treated as a manual/process step. This keeps the
# runner away from destructive or environment-mutating actions (PyPI publish,
# `uv tool install`, creating new-Python venvs, etc.).
SAFE_PREFIXES = (
    "potpie --version",
    "potpie status",
    "potpie doctor",
    "potpie whoami",
    "potpie setup",
    "potpie source add",
    "potpie source list",
    "potpie resolve",
    "potpie search",
    "potpie config get",
    "potpie graph",
)

BACKTICK_RE = re.compile(r"`([^`]+)`")
MAX_EVIDENCE_CHARS = 6000

# Some safe commands require a positional argument the test plan omits (it lists
# them as bare `potpie resolve` / `potpie search`). Supply a real default so the
# command actually exercises the feature instead of erroring on a missing arg.
# Keys are matched against the bare command (no trailing args); values are
# overridable from the CLI (--resolve-task / --search-query).
DEFAULT_ARGS = {
    "potpie resolve": 'How does the CLI host routing work?',
    "potpie search": 'host_cli',
}


# --------------------------------------------------------------------------- #
# Result container
# --------------------------------------------------------------------------- #
class TestResult:
    def __init__(self, row):
        self.seq = row.get("Seq")
        self.tid = row.get("ID")
        self.type = row.get("Type")
        self.workstream = row.get("Workstream")
        self.priority = row.get("Priority")
        self.step = row.get("Step / Command") or ""
        self.expected = row.get("Expected Result / Exit Criteria") or ""
        self.commands_run = []
        self.result = "PENDING"
        self.exit_code = ""
        self.evidence = ""
        self.timestamp = ""


class IsolatedRun:
    """Owns per-test temp state so CLI calls never touch the user's Potpie home."""

    def __init__(self, root: Path, repo: str):
        self.root = root
        self.repo = Path(repo)
        self._counter = 0
        self._flows: dict[str, dict] = {}

    def context(self, isolation: str | None) -> dict:
        mode = isolation or "shared"
        if mode == "none":
            return {"cwd": self.repo, "env": os.environ.copy(), "state": {}}
        if mode.startswith("flow:"):
            name = mode.split(":", 1)[1]
            if name not in self._flows:
                self._flows[name] = self._make_context(name)
            return self._flows[name]
        return self._make_context("case")

    def _make_context(self, label: str) -> dict:
        self._counter += 1
        base = self.root / f"{self._counter:03d}-{label}"
        home = base / "home"
        xdg_config = base / "xdg-config"
        xdg_cache = base / "xdg-cache"
        xdg_data = base / "xdg-data"
        xdg_state = base / "xdg-state"
        ce_home = base / "context-engine-home"
        repo = base / "repo"
        for path in (home, xdg_config, xdg_cache, xdg_data, xdg_state, ce_home, repo):
            path.mkdir(parents=True, exist_ok=True)
        (repo / ".git").mkdir(exist_ok=True)
        (repo / "README.md").write_text("# isolated context-engine cli test\n", encoding="utf-8")
        (repo / "app.py").write_text("def hello():\n    return 'world'\n", encoding="utf-8")
        env = os.environ.copy()
        env.update(
            {
                "HOME": str(home),
                "XDG_CONFIG_HOME": str(xdg_config),
                "XDG_CACHE_HOME": str(xdg_cache),
                "XDG_DATA_HOME": str(xdg_data),
                "XDG_STATE_HOME": str(xdg_state),
                "UV_CACHE_DIR": str(xdg_cache / "uv"),
                "CONTEXT_ENGINE_HOME": str(ce_home),
                "CONTEXT_ENGINE_HOST_MODE": "in_process",
                "CONTEXT_ENGINE_BACKEND": "embedded",
                "CONTEXT_ENGINE_EMBEDDER": "none",
                "PYTHON_KEYRING_BACKEND": "keyring.backends.null.Keyring",
                "POTPIE_DISABLE_TELEMETRY": "1",
            }
        )
        return {
            "cwd": repo,
            "env": env,
            "state": {
                "base": str(base),
                "home": str(home),
                "ce_home": str(ce_home),
                "repo": str(repo),
                "mutation_file": str(base / "mutation.json"),
                "bulk_file": str(base / "bulk.json"),
                "manifest_file": str(base / "bulk-manifest.json"),
                "pot_name": f"qa-pot-{uuid.uuid4().hex[:8]}",
                "second_pot_name": f"qa-pot-{uuid.uuid4().hex[:8]}",
            },
        }


# --------------------------------------------------------------------------- #
# Parsing / classification
# --------------------------------------------------------------------------- #
def extract_commands(step_text: str) -> list[str]:
    """Return backtick-quoted command-looking snippets from a step description."""
    out = []
    for raw in BACKTICK_RE.findall(step_text or ""):
        cmd = raw.strip()
        if cmd:
            out.append(cmd)
    return out


def is_runnable(cmd: str) -> bool:
    return any(cmd.startswith(p) for p in SAFE_PREFIXES)


def augment_command(cmd: str, default_args: dict[str, str]) -> str:
    """Append a default positional argument to bare arg-requiring commands.

    e.g. ``potpie resolve`` -> ``potpie resolve "How does ... work?"`` so the
    test exercises the command instead of failing on a missing argument. A
    command that already carries its own argument is left untouched.
    """
    bare = cmd.strip()
    if bare in default_args:
        return f'{bare} "{default_args[bare]}"'
    return cmd


def classify(commands: list[str]) -> tuple[list[str], str]:
    """Return (runnable_commands, manual_reason).

    manual_reason is empty when the test is auto-runnable.
    """
    runnable = [c for c in commands if is_runnable(c)]
    if runnable:
        return runnable, ""
    if commands:
        return [], "Manual/process step (commands present but not auto-runnable: %s)" % "; ".join(commands)
    return [], "Manual/process step (no executable CLI command)"


def deterministic_context_engine_tests() -> list[dict]:
    """Rows appended to the workbook plan for deterministic local CLI coverage."""
    return [
        _det_row("CE-ENTRY-01", "Entrypoint", "Run `{cli} --version`.", "Version command exits cleanly."),
        _det_row("CE-ENTRY-02", "Entrypoint", "Run `{cli} --help`.", "Help renders without importing user state."),
        _det_row("CE-ISO-01", "Isolation", "Run `{cli} --json pot create leaked-pot --use`.", "A pot can be created in one isolated home.", isolation="isolated"),
        _det_row("CE-ISO-02", "Isolation", "Run `{cli} --json pot use leaked-pot`.", "A fresh isolated home cannot see the pot from CE-ISO-01.", isolation="isolated", expected_exit=1),
        _det_row("CE-ISO-03", "Isolation", "Run `{cli} --json pot list`.", "Fresh isolated pot list does not inherit CE-ISO-01 state.", isolation="isolated", checks="no_leaked_pot"),
        _det_row("CE-SETUP-01", "Setup", "Run `{cli} --json setup --repo . --agent claude --backend embedded --yes --in-process`.", "Setup completes in an isolated repo/home.", isolation="flow:setup"),
        _det_row("CE-SETUP-02", "Setup", "Run `{cli} --json setup --repo . --agent claude --backend embedded --yes --in-process`.", "Setup is idempotent in the same isolated flow.", isolation="flow:setup"),
        _det_row("CE-FILES-01", "Local storage", "Verify local context-engine config, storage, and log directories.", "Only the isolated context-engine home is used.", isolation="flow:setup", checks="local_storage"),
        _det_row("CE-SKILL-01", "Local skill files", "Run `{cli} --json skills install --agent claude --scope project --path .`.", "Claude skill files are installed into the isolated repo.", isolation="flow:setup"),
        _det_row("CE-SKILL-02", "Local skill files", "Verify CLAUDE.md and local skill markers.", "Skill file exists and contains Potpie-managed instructions.", isolation="flow:setup", checks="claude_skill_file"),
        _det_row("CE-CONFIG-01", "Config", "Run `{cli} --json config set qa.runner isolated`.", "Config set writes to isolated config.json.", isolation="flow:setup"),
        _det_row("CE-CONFIG-02", "Config", "Run `{cli} --json config get qa.runner`.", "Config get reads the isolated value.", isolation="flow:setup"),
        _det_row("CE-POT-01", "Pot management", "Run `{cli} --json pot create {pot_name} --use`.", "Creates an active local pot in the flow.", isolation="flow:graph", capture="pot_id=id"),
        _det_row("CE-POT-02", "Pot management", "Run `{cli} --json pot create {second_pot_name}`.", "Creates a second local pot without touching user state.", isolation="flow:graph", capture="second_pot_id=id"),
        _det_row("CE-POT-03", "Pot management", "Run `{cli} --json pot list`.", "Lists only flow-local pots.", isolation="flow:graph"),
        _det_row("CE-POT-04", "Pot management", "Run `{cli} --json pot info`.", "Active pot details are available.", isolation="flow:graph"),
        _det_row("CE-POT-05", "Pot management", "Run `{cli} --json pot linked --repo current`.", "Repo linked-pot lookup is deterministic.", isolation="flow:graph"),
        _det_row("CE-POT-06", "Pot management", "Run `{cli} --json pot default show --repo current`.", "Repo default lookup is deterministic.", isolation="flow:graph"),
        _det_row("CE-POT-07", "Pot management", "Run `{cli} --json pot default set {pot_id} --repo current`.", "Repo default can be set to the captured pot.", isolation="flow:graph"),
        _det_row("CE-POT-08", "Pot management", "Run `{cli} --json pot default clear --repo current`.", "Repo default can be cleared in isolated state.", isolation="flow:graph"),
        _det_row("CE-SOURCE-01", "Source registry", "Run `{cli} --json source add repo . --pot {pot_id}`.", "Registers the isolated repo as a source.", isolation="flow:graph", capture="source_id=source_id"),
        _det_row("CE-SOURCE-02", "Source registry", "Run `{cli} --json source list --pot {pot_id}`.", "Source list includes the isolated source.", isolation="flow:graph"),
        _det_row("CE-SOURCE-03", "Source registry", "Run `{cli} --json source status {source_id} --pot {pot_id}`.", "Captured source status is readable.", isolation="flow:graph"),
        _det_row("CE-BACKEND-01", "Backend", "Run `{cli} --json backend list`.", "Backend profiles render.", isolation="flow:graph"),
        _det_row("CE-BACKEND-02", "Backend", "Run `{cli} --json backend status`.", "Active backend status renders.", isolation="flow:graph"),
        _det_row("CE-BACKEND-03", "Backend", "Run `{cli} --json backend doctor`.", "Backend doctor returns readiness information.", isolation="flow:graph"),
        _det_row("CE-GRAPH-01", "Graph discovery", "Run `{cli} --json graph status --pot {pot_id}`.", "Graph status returns a structured envelope.", isolation="flow:graph"),
        _det_row("CE-GRAPH-02", "Graph discovery", "Run `{cli} --json graph catalog --task deterministic-cli-test --pot {pot_id}`.", "Graph catalog ranks views.", isolation="flow:graph"),
        _det_row("CE-GRAPH-03", "Graph discovery", "Run `{cli} --json graph describe infra_topology --view service_neighborhood --examples`.", "Graph describe returns the view contract.", isolation="flow:graph"),
        _det_row("CE-GRAPH-04", "Graph write", "Write deterministic graph mutation fixture.", "Mutation fixture exists inside isolated temp state.", isolation="flow:graph", checks="write_mutation_fixture"),
        _det_row("CE-GRAPH-05", "Graph write", "Run `{cli} --json graph mutate --file {mutation_file} --pot {pot_id}`.", "Graph mutation applies to the isolated pot.", isolation="flow:graph"),
        _det_row("CE-FILES-02", "Local storage", "Verify graph backend files are isolated.", "Graph backend files exist only under isolated context-engine home.", isolation="flow:graph", checks="backend_storage"),
        _det_row("CE-GRAPH-06", "Graph read", "Run `{cli} --json graph read --subgraph infra_topology --view service_neighborhood --scope service:journey-service --limit 10 --pot {pot_id}`.", "Graph read returns the synthetic service context.", isolation="flow:graph"),
        _det_row("CE-GRAPH-07", "Graph read", "Run `{cli} --json graph search-entities journey-service --pot {pot_id}`.", "Entity search finds the synthetic service.", isolation="flow:graph"),
        _det_row("CE-GRAPH-08", "Graph read", "Run `{cli} --json graph neighborhood --entity service:journey-service --predicate DEPENDS_ON --depth 1 --pot {pot_id}`.", "Neighborhood reads the synthetic dependency edge.", isolation="flow:graph"),
        _det_row("CE-GRAPH-09", "Graph history", "Run `{cli} --json graph history --entity service:journey-service --limit 10 --pot {pot_id}`.", "History command returns a structured response.", isolation="flow:graph"),
        _det_row("CE-GRAPH-10", "Graph write", "Run `{cli} --json graph propose --file {mutation_file} --pot {pot_id}`.", "Graph propose validates the deterministic mutation.", isolation="flow:graph", capture="plan_id=plan_id"),
        _det_row("CE-GRAPH-11", "Graph write", "Run `{cli} --json graph commit {plan_id} --verify --pot {pot_id}`.", "Graph commit applies the captured plan.", isolation="flow:graph"),
        _det_row("CE-BULK-01", "Graph bulk", "Write deterministic bulk mutation fixture.", "Bulk fixture exists inside isolated temp state.", isolation="flow:graph", checks="write_bulk_fixture"),
        _det_row("CE-BULK-02", "Graph bulk", "Run `{cli} --json graph bulk apply --file {bulk_file} --chunk-size 1 --dry-run --manifest {manifest_file} --pot {pot_id}`.", "Bulk dry-run validates chunks and writes a manifest.", isolation="flow:graph", checks="bulk_manifest"),
        _det_row("CE-INBOX-01", "Graph inbox", "Run `{cli} --json graph inbox add --summary deterministic-cli-inbox --evidence test:run-plan --subgraph infra_topology --created-by qa-runner --pot {pot_id}`.", "Inbox item is created in isolated graph state.", isolation="flow:graph", capture="inbox_id=item.item_id"),
        _det_row("CE-INBOX-02", "Graph inbox", "Run `{cli} --json graph inbox list --limit 10 --pot {pot_id}`.", "Inbox list is structured.", isolation="flow:graph"),
        _det_row("CE-INBOX-03", "Graph inbox", "Run `{cli} --json graph inbox show {inbox_id} --pot {pot_id}`.", "Captured inbox item can be shown.", isolation="flow:graph"),
        _det_row("CE-INBOX-04", "Graph inbox", "Run `{cli} --json graph inbox claim {inbox_id} --by qa-runner --pot {pot_id}`.", "Captured inbox item can be claimed.", isolation="flow:graph"),
        _det_row("CE-INBOX-05", "Graph inbox", "Run `{cli} --json graph inbox mark-rejected {inbox_id} --reason deterministic-close --by qa-runner --pot {pot_id}`.", "Captured inbox item can be closed rejected.", isolation="flow:graph"),
        _det_row("CE-QUALITY-01", "Graph quality", "Run `{cli} --json graph quality summary --pot {pot_id}`.", "Quality summary returns structured output.", isolation="flow:graph"),
        _det_row("CE-QUALITY-02", "Graph quality", "Run `{cli} --json graph quality duplicate-candidates --limit 5 --pot {pot_id}`.", "Duplicate-candidates report is structured.", isolation="flow:graph"),
        _det_row("CE-QUALITY-03", "Graph quality", "Run `{cli} --json graph quality stale-facts --limit 5 --pot {pot_id}`.", "Stale-facts report is structured.", isolation="flow:graph"),
        _det_row("CE-QUALITY-04", "Graph quality", "Run `{cli} --json graph quality conflicting-claims --limit 5 --pot {pot_id}`.", "Conflicting-claims report is structured.", isolation="flow:graph"),
        _det_row("CE-QUALITY-05", "Graph quality", "Run `{cli} --json graph quality orphan-entities --limit 5 --pot {pot_id}`.", "Orphan-entities report is structured.", isolation="flow:graph"),
        _det_row("CE-QUALITY-06", "Graph quality", "Run `{cli} --json graph quality low-confidence --limit 5 --threshold 0.5 --pot {pot_id}`.", "Low-confidence report is structured.", isolation="flow:graph"),
        _det_row("CE-QUALITY-07", "Graph quality", "Run `{cli} --json graph quality projection-drift --limit 5 --pot {pot_id}`.", "Projection-drift report is structured.", isolation="flow:graph"),
        _det_row("CE-QUERY-01", "Query", "Run `{cli} --json resolve deterministic-cli-test --pot {pot_id}`.", "Resolve returns structured context or clear empty-context output.", isolation="flow:graph"),
        _det_row("CE-QUERY-02", "Query", "Run `{cli} --json search journey-service --pot {pot_id}`.", "Search command is deterministic against the isolated pot.", isolation="flow:graph"),
        _det_row("CE-DAEMON-01", "Daemon", "Run `{cli} --json daemon status`.", "Daemon status uses isolated context-engine home.", isolation="flow:daemon"),
        _det_row("CE-DAEMON-02", "Daemon", "Run `{cli} --json daemon logs`.", "Daemon logs command is safe when no log exists.", isolation="flow:daemon"),
        _det_row("CE-SERVICE-01", "Managed services", "Run `{cli} --json service status`.", "Service status fails cleanly when no detached daemon is running.", isolation="flow:daemon", expected_exit=2),
        _det_row("CE-ERROR-01", "Error handling", "Run `{cli} --json graph describe not_a_subgraph`.", "Invalid graph describe returns structured validation failure.", isolation="isolated", expected_exit=1),
        _det_row("CE-ERROR-02", "Error handling", "Run `{cli} --json graph mutate --file missing.json`.", "Missing mutation file returns a failure without touching user state.", isolation="isolated", expected_exit=1),
    ]


def _det_row(
    tid: str,
    workstream: str,
    step: str,
    expected: str,
    *,
    isolation: str = "isolated",
    expected_exit: int = 0,
    capture: str = "",
    checks: str = "",
) -> dict:
    return {
        "Seq": "",
        "ID": tid,
        "Type": "Deterministic",
        "Workstream": workstream,
        "Priority": "P0",
        "Step / Command": step,
        "Expected Result / Exit Criteria": expected,
        "_deterministic": True,
        "_isolation": isolation,
        "_expected_exit": expected_exit,
        "_capture": capture,
        "_checks": checks,
    }


def default_cli_cmd(repo: str) -> str:
    project = Path(repo).expanduser().resolve() / "potpie" / "context-engine"
    return f"uv run --project {shlex.quote(str(project))} --no-sync potpie"


# --------------------------------------------------------------------------- #
# Execution
# --------------------------------------------------------------------------- #
def run_command(
    cmd: str,
    cwd: str | Path,
    timeout: int,
    *,
    env: dict[str, str] | None = None,
) -> tuple[int, str]:
    try:
        proc = subprocess.run(  # noqa: S603 - plan commands are allow-listed or built-in deterministic specs.
            shlex.split(cmd),
            cwd=cwd,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=env,
        )
        out = (proc.stdout or "") + (proc.stderr or "")
        return proc.returncode, out.strip()
    except subprocess.TimeoutExpired as exc:
        partial = (exc.stdout or "") if isinstance(exc.stdout, str) else ""
        return 124, f"TIMEOUT after {timeout}s\n{partial}".strip()
    except Exception as exc:  # pragma: no cover
        return 1, f"RUNNER ERROR: {exc}"


def git_checkout(repo: str, branch: str) -> tuple[bool, str]:
    code, out = run_command("git rev-parse --abbrev-ref HEAD", repo, 30)
    current = out.strip()
    if current == branch:
        return True, f"Already on '{branch}'."
    code, out = run_command(f"git checkout {branch}", repo, 60)
    if code != 0:
        return False, out
    code2, head = run_command("git rev-parse --abbrev-ref HEAD", repo, 30)
    return True, f"Switched from '{current}' to '{head.strip()}'.\n{out}".strip()


class _FormatState(dict):
    def __missing__(self, key):
        return "{" + key + "}"


def render_command(cmd: str, state: dict[str, str], cli_cmd: str) -> str:
    values = _FormatState(state)
    values["cli"] = cli_cmd
    return cmd.format_map(values)


def unwrap_json_payload(output: str) -> dict:
    try:
        payload = json.loads(output or "{}")
    except json.JSONDecodeError:
        return {}
    result = payload.get("result")
    if isinstance(result, dict):
        return result
    return payload if isinstance(payload, dict) else {}


def capture_values(spec: str, output: str, state: dict[str, str]) -> None:
    if not spec:
        return
    payload = unwrap_json_payload(output)
    for part in spec.split(","):
        if not part.strip():
            continue
        target, _, path = part.partition("=")
        target = target.strip()
        path = (path or target).strip()
        value = _payload_path(payload, path)
        if value is not None:
            state[target] = str(value)


def _payload_path(payload: dict, path: str):
    cur = payload
    for key in path.split("."):
        if isinstance(cur, dict) and key in cur:
            cur = cur[key]
        else:
            return None
    return cur


def run_post_check(check: str, ctx: dict, last_output: str = "") -> tuple[int, str]:
    if not check:
        return 0, ""
    state = ctx["state"]
    cwd = Path(ctx["cwd"])
    ce_home = Path(state["ce_home"])
    if check == "write_mutation_fixture":
        _write_mutation_fixture(Path(state["mutation_file"]))
        return 0, f"wrote {state['mutation_file']}"
    if check == "write_bulk_fixture":
        _write_bulk_fixture(Path(state["bulk_file"]))
        return 0, f"wrote {state['bulk_file']}"
    if check == "bulk_manifest":
        manifest = Path(state["manifest_file"])
        if not manifest.exists():
            return 1, f"missing bulk manifest: {manifest}"
        return 0, f"bulk manifest exists: {manifest}"
    if check == "local_storage":
        required = [ce_home, ce_home / "config.json"]
        missing = [str(p) for p in required if not p.exists()]
        if missing:
            return 1, "missing local storage paths: " + ", ".join(missing)
        return 0, f"local storage isolated under {ce_home}"
    if check == "backend_storage":
        expected = ce_home / "graph.json"
        if expected.exists():
            return 0, f"backend storage isolated under {expected}"
        files = [str(p.relative_to(ce_home)) for p in ce_home.rglob("*") if p.is_file()]
        if not files:
            return 1, f"no backend/storage files found under isolated home {ce_home}"
        return 0, "backend/local files under isolated home: " + ", ".join(files[:12])
    if check == "claude_skill_file":
        claude = cwd / "CLAUDE.md"
        if not claude.exists():
            return 1, f"missing local skill file: {claude}"
        text = claude.read_text(encoding="utf-8", errors="replace")
        if "potpie-start" not in text and "Potpie" not in text:
            return 1, f"CLAUDE.md does not contain Potpie skill markers: {claude}"
        return 0, f"verified local skill file: {claude}"
    if check == "no_leaked_pot":
        if "leaked-pot" in last_output:
            return 1, "fresh isolated pot list unexpectedly contains leaked-pot"
        return 0, "fresh isolated state does not contain leaked-pot"
    return 1, f"unknown post-check: {check}"


def _write_mutation_fixture(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "operations": [
                    {
                        "op": "upsert_entity",
                        "subject": {
                            "key": "service:journey-service",
                            "type": "Service",
                            "name": "journey-service",
                            "summary": "Service used by deterministic CLI runner tests.",
                            "description": "Synthetic service entity for local context-engine CLI validation.",
                        },
                    },
                    {
                        "op": "upsert_entity",
                        "subject": {
                            "key": "service:journey-ledger",
                            "type": "Service",
                            "name": "journey-ledger",
                            "summary": "Dependency used by deterministic CLI runner tests.",
                            "description": "Synthetic dependency entity for local context-engine CLI validation.",
                        },
                    },
                    {
                        "op": "link_entities",
                        "subgraph": "infra_topology",
                        "subject": {"key": "service:journey-service", "type": "Service"},
                        "predicate": "DEPENDS_ON",
                        "object": {"key": "service:journey-ledger", "type": "Service"},
                        "truth": "source_observation",
                        "description": "journey service depends on journey ledger",
                        "evidence": [
                            {
                                "source_ref": "test:deterministic-cli-runner",
                                "authority": "repository_metadata",
                            }
                        ],
                    },
                ]
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _write_bulk_fixture(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "operations": [
                    {
                        "op": "upsert_entity",
                        "subject": {
                            "key": "feature:deterministic-cli-bulk",
                            "type": "Feature",
                            "name": "deterministic-cli-bulk",
                            "summary": "Synthetic feature used by bulk CLI validation.",
                        },
                    }
                ]
            },
            indent=2,
        ),
        encoding="utf-8",
    )


# --------------------------------------------------------------------------- #
# Workbook IO
# --------------------------------------------------------------------------- #
def load_tests(wb) -> list[dict]:
    ws = wb[SOURCE_SHEET]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            headers[c] = str(v).strip()
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row=r, column=4).value is None:  # column D = ID
            continue
        row = {headers[c]: ws.cell(row=r, column=c).value for c in headers}
        rows.append(row)
    return rows


def unique_sheet_name(wb, base: str) -> str:
    name = base[:31]
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        candidate = f"{base[:28]}_{i}"
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


RESULT_FILLS = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
    "MANUAL": PatternFill("solid", fgColor="FFEB9C"),
    "SKIPPED": PatternFill("solid", fgColor="D9D9D9"),
}


def write_results(wb, sheet_name: str, results: list[TestResult], meta: dict):
    ws = wb.create_sheet(title=sheet_name)
    cols = [
        ("Seq", 6),
        ("ID", 26),
        ("Type", 12),
        ("Workstream", 20),
        ("Priority", 8),
        ("Step / Command", 50),
        ("Expected Result", 45),
        ("Command(s) Run", 38),
        ("Result", 10),
        ("Exit Code", 9),
        ("Evidence", 70),
        ("Run Timestamp", 20),
    ]

    # Metadata banner
    ws.cell(row=1, column=1, value=f"Test run: {meta['run_at']}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Branch: {meta['branch']}  |  Repo: {meta['repo']}")
    ws.cell(row=3, column=1, value=f"Checkout: {meta['checkout']}")
    summary = "  ".join(f"{k}={v}" for k, v in meta["summary"].items())
    ws.cell(row=4, column=1, value=f"Summary: {summary}").font = Font(bold=True)

    header_row = 6
    header_fill = PatternFill("solid", fgColor="305496")
    for ci, (title, width) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=ci, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, res in enumerate(results, start=header_row + 1):
        values = [
            res.seq,
            res.tid,
            res.type,
            res.workstream,
            res.priority,
            res.step,
            res.expected,
            "\n".join(res.commands_run),
            res.result,
            res.exit_code,
            res.evidence,
            res.timestamp,
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        rcell = ws.cell(row=ri, column=9)
        rcell.font = Font(bold=True)
        if res.result in RESULT_FILLS:
            rcell.fill = RESULT_FILLS[res.result]

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to the workbook.")
    ap.add_argument("--repo", default=DEFAULT_REPO, help="Repo to run CLI tests in.")
    ap.add_argument("--branch", default="main", help="Branch to switch to before testing.")
    ap.add_argument("--no-checkout", action="store_true", help="Do not change git branch.")
    ap.add_argument("--dry-run", action="store_true", help="Classify only; execute nothing.")
    ap.add_argument("--timeout", type=int, default=240, help="Per-command timeout (seconds).")
    ap.add_argument("--sheet", default=None, help="Override the result sheet name.")
    ap.add_argument(
        "--cli-cmd",
        default=LOCAL_CLI,
        help="Command prefix used for deterministic local CLI rows.",
    )
    ap.add_argument(
        "--no-deterministic-context-engine",
        action="store_true",
        help="Do not append built-in deterministic context-engine CLI coverage rows.",
    )
    ap.add_argument(
        "--no-isolation",
        action="store_true",
        help="Run workbook auto-runnable commands in the provided repo/user environment.",
    )
    ap.add_argument(
        "--only-deterministic",
        action="store_true",
        help="Run only the built-in deterministic context-engine CLI rows.",
    )
    ap.add_argument(
        "--only-id",
        default=None,
        help="Run rows whose ID exactly matches or starts with this value.",
    )
    ap.add_argument("--resolve-task", default=DEFAULT_ARGS["potpie resolve"],
                    help="Default TASK argument injected into bare `potpie resolve`.")
    ap.add_argument("--search-query", default=DEFAULT_ARGS["potpie search"],
                    help="Default QUERY argument injected into bare `potpie search`.")
    ap.add_argument("--keep-existing", action="store_true",
                    help="Keep an existing same-date sheet and write a suffixed one instead of replacing it.")
    args = ap.parse_args(argv)
    if not args.cli_cmd:
        args.cli_cmd = default_cli_cmd(args.repo)

    default_args = {
        "potpie resolve": args.resolve_task,
        "potpie search": args.search_query,
    }

    if not os.path.exists(args.excel):
        sys.exit(f"Workbook not found: {args.excel}")

    run_at = dt.datetime.now()
    run_stamp = run_at.strftime("%Y-%m-%d %H:%M:%S")

    # 1. Branch checkout
    checkout_msg = "skipped (--no-checkout)" if args.no_checkout else "skipped (dry-run)"
    if not args.no_checkout and not args.dry_run:
        ok, checkout_msg = git_checkout(args.repo, args.branch)
        print(f"[git] {checkout_msg}")
        if not ok:
            sys.exit(f"Failed to checkout '{args.branch}': {checkout_msg}")
    else:
        print(f"[git] checkout {checkout_msg}")

    # 2. Load workbook + tests
    wb = openpyxl.load_workbook(args.excel)
    wb_data = openpyxl.load_workbook(args.excel, data_only=True)  # for cached values
    tests = load_tests(wb_data)
    if not args.no_deterministic_context_engine:
        tests.extend(deterministic_context_engine_tests())
    if args.only_deterministic:
        tests = [row for row in tests if row.get("_deterministic")]
    if args.only_id:
        tests = [
            row
            for row in tests
            if str(row.get("ID") or "") == args.only_id
            or str(row.get("ID") or "").startswith(args.only_id)
        ]
    print(f"[plan] {len(tests)} tests loaded from '{SOURCE_SHEET}'\n")

    results: list[TestResult] = []
    counts = {"PASS": 0, "FAIL": 0, "MANUAL": 0, "SKIPPED": 0}
    isolation_root_mgr = tempfile.TemporaryDirectory(prefix="potpie-cli-plan-")
    isolation = IsolatedRun(Path(isolation_root_mgr.name), args.repo)

    for i, row in enumerate(tests, start=1):
        res = TestResult(row)
        commands = extract_commands(res.step)
        deterministic = bool(row.get("_deterministic"))
        runnable, manual_reason = (commands, "") if deterministic else classify(commands)
        label = f"[{i}/{len(tests)}] {res.tid}"

        if not runnable and not row.get("_checks"):
            res.result = "MANUAL"
            res.evidence = manual_reason
            res.timestamp = run_stamp
            counts["MANUAL"] += 1
            print(f"{label}  MANUAL  - {manual_reason[:80]}")
            results.append(res)
            continue

        run_ctx = isolation.context(
            row.get("_isolation") if deterministic else ("none" if args.no_isolation else "isolated")
        )
        run_state = run_ctx["state"]
        runnable = [
            render_command(augment_command(c, default_args), run_state, args.cli_cmd)
            for c in runnable
        ]
        res.commands_run = runnable
        if args.dry_run:
            res.result = "SKIPPED"
            res.evidence = (
                f"dry-run: isolation={row.get('_isolation') or ('none' if args.no_isolation else 'isolated')}"
                " would execute -> "
                + " ; ".join(runnable)
            )
            res.timestamp = run_stamp
            counts["SKIPPED"] += 1
            print(f"{label}  DRY     - would run: {' ; '.join(runnable)}")
            results.append(res)
            continue

        # Execute each runnable command, collect evidence
        all_ok = True
        evidence_chunks = []
        last_code = 0
        last_output = ""
        expected_exit = int(row.get("_expected_exit") or 0)
        for cmd in runnable:
            print(f"{label}  RUN     $ {cmd}")
            code, out = run_command(
                cmd,
                run_ctx["cwd"],
                args.timeout,
                env=run_ctx["env"],
            )
            last_code = code
            last_output = out
            status = "ok" if code == expected_exit else f"exit {code}"
            evidence_chunks.append(f"$ {cmd}  ->  {status}\n{out}")
            if code != expected_exit:
                all_ok = False
            elif deterministic:
                capture_values(str(row.get("_capture") or ""), out, run_state)
        check = str(row.get("_checks") or "")
        if check and not args.dry_run:
            code, out = run_post_check(check, run_ctx, last_output)
            last_code = code
            status = "ok" if code == 0 else f"exit {code}"
            evidence_chunks.append(f"[check:{check}]  ->  {status}\n{out}")
            if code != 0:
                all_ok = False
        evidence = "\n\n".join(evidence_chunks).strip()
        if len(evidence) > MAX_EVIDENCE_CHARS:
            evidence = evidence[:MAX_EVIDENCE_CHARS] + "\n... [truncated]"

        res.result = "PASS" if all_ok else "FAIL"
        res.exit_code = last_code
        res.evidence = evidence
        res.timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        counts[res.result] += 1
        print(f"{label}  {res.result}")
        results.append(res)

    # 3. Write results sheet. Re-running on the same date refreshes that day's
    # sheet by default; pass --keep-existing to append a suffixed sheet instead.
    if args.sheet:
        sheet_name = args.sheet
    else:
        base = run_at.strftime("%Y-%m-%d")
        if args.keep_existing:
            sheet_name = unique_sheet_name(wb, base)
        else:
            sheet_name = base[:31]
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
    meta = {
        "run_at": run_stamp,
        "branch": "(not changed)" if (args.no_checkout or args.dry_run) else args.branch,
        "repo": args.repo,
        "checkout": checkout_msg,
        "summary": counts,
    }
    write_results(wb, sheet_name, results, meta)
    if args.dry_run:
        print("\n[dry-run] workbook NOT modified.")
    else:
        wb.save(args.excel)

    print("\n" + "=" * 60)
    if not args.dry_run:
        print(f"Results written to sheet '{sheet_name}' in {args.excel}")
    print("Summary: " + "  ".join(f"{k}={v}" for k, v in counts.items()))
    print("=" * 60)


if __name__ == "__main__":
    main()
